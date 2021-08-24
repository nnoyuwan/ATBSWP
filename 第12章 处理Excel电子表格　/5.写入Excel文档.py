# OpenPyXL 也提供了一些方法写入数据，这意味着你的程序可以创建和编辑电子
# 表格文件。利用 Python，创建一个包含几千行数据的电子表格是非常简单的

# 12.5.1 创建并保存 Excel 文档
import openpyxl

# """Workbook is the container for all other parts of the document."""
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)

# 当修改 Workbook 对象或它的工作表和单元格时，电子表格文件不会保存，除非你调
# 用 save()工作簿方法。
wb.save('example_copy.xlsx')

# 当你编辑从文件中加载的一个电子表格时，总是应该将新的、编辑过的电子表
# 格保存到不同的文件名中。这样，如果代码中有缺陷，导致新的保存到文件中数据
# 不对或讹误，还有最初的电子表格文件可以处理。

# 12.5.2 创建和删除工作表

wb = openpyxl.Workbook()
print(wb.sheetnames)
wb.create_sheet()  # <Worksheet "Sheet1">
print(wb.sheetnames)
wb.create_sheet(title='First Sheet', index=0)
print(wb.sheetnames)  # ['First Sheet', 'Sheet', 'Sheet1']
wb.create_sheet(title='Middle Sheet', index=2)
print(wb.sheetnames)  # ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']

wb.remove_sheet(wb['Middle Sheet'])
wb.remove_sheet(wb['Sheet1'])
print(wb.sheetnames)  # ['First Sheet', 'Sheet']

# remove_sheet()方法接受一个 Worksheet 对象作为其参数，而不是工作表名称的字符
# 串。如果你只知道要删除的工作表的名称，就调用 get_sheet_by_name()，将它的返
# 回值传入 remove_sheet()。
# 在工作簿中添加或删除工作表之后，记得调用 save()方法来保存变更。


# 12.5.3 将值写入单元格

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello World!'
print(sheet['A1'].value)

# 如果你有单元格坐标的字符串(类似A1,B2这种)，可以像字典的键一样，将它用于 Worksheet 对
# 象，指定要写入的单元格。
