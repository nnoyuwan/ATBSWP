# 公式以一个等号开始，可以配置单元格，让它包含通过其他单元格计算得到的
# 值。在本节中，你将利用 openpyxl 模块，用编程的方式在单元格中添加公式，就像
# 添加普通的值一样。例如：


# 例如：
# >>> sheet['B9'] = '=SUM(B1:B8)'
# 这将=SUM(B1:B8)作为单元格 B9 的值。这将 B9 单元格设置为一个公式，计算
# 单元格 B1 到 B8 的和。
import os

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=sum(A1:A2)'
wb.save('writeFormula.xlsx')

# 也可以读取单元格中的公式，就像其他值一样。但是，如果你希望看到该公式
# 的计算结果，而不是原来的公式，就必须将 load_workbook()的 data_only 关键字参
# 数设置为 True。这意味着 Workbook 对象要么显示公式，要么显示公式的结果，不
# 能兼得

import os

os.getcwd()
os.chdir('C:\\Users\\admin\\PycharmProjects\\ATBSWP\\第12章 处理Excel电子表格　')
wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wbFormulas.active
print(sheet['A3'].value)

wbFormulas = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
sheet = wbFormulas.active
print(sheet['A3'].value)
