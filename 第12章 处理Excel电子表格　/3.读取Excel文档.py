# 本章的例子将使用一个电子表格 example.xlsx，它保存在根文件夹中。
# 。图 12-1 展示了 3 个默认的表，名为 Sheet1、 Sheet2 和 Sheet3，这是 Excel 自动为新工作簿提供的
# （不同操作系统和电子表格程序，提供的默认表个数可能会不同）

# 既然有了示例电子表格，就来看看如何用 openpyxl 模块来操作它。

# 表 12-1 example.xlsx 电子表格
#           A                   B         C
# 1 4/5/2015 1:34:02 PM     Apples          73
# 2 4/5/2015 3:41:23 AM     Cherries        85
# 3 4/6/2015 12:46:51 PM    Pears           14
# 4 4/8/2015 8:59:43 AM     Oranges         52
# 5 4/10/2015 2:07:00 AM    Apples          152
# 6 4/10/2015 6:10:37 PM    Bananas         23
# 7 4/10/2015 2:40:46 AM    Strawberries    98

# 12.3.1 用 openpyxl 模块打开 Excel 文档
import os

import openpyxl

os.getcwd()
os.chdir('第12章 处理Excel电子表格　')
wb = openpyxl.load_workbook('example.xlsx')
type(wb)

# openpyxl.load_workbook()函数接受文件名，返回一个 workbook 数据类型的值。这
# 个 workbook 对象代表这个 Excel 文件，有点类似 File 对象代表一个打开的文本文件。
# 要记住， example.xlsx 需要在当前工作目录，你才能处理它。可以导入 os，使
# 用函数 os.getcwd()弄清楚当前工作目录是什么，并使用 os.chdir()改变当前工作目录。

# 12.3.2 从工作簿中取得工作表
wb.get_sheet_names()  # DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames).
sheetnames = wb.sheetnames  # ['Sheet1', 'Sheet2', 'Sheet3']

sheet = wb.get_sheet_by_name(
    'Sheet1')  # DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
sheet = wb['Sheet1']
sheet
type(sheet)  # <class 'openpyxl.worksheet.worksheet.Worksheet'>

print(sheet.title)  # 'Sheet1'

# anotherSheet = wb.get_active_sheet()  # 'Workbook' object has no attribute 'get_activate_sheet'
anotherSheet = wb.active
anotherSheet  # <Worksheet "Sheet1">

# 递表名字符串获得。最后，可以调用 Workbook 对象的 get_active_sheet()方法，取得
# 工作簿的活动表。活动表是工作簿在 Excel 中打开时出现的工作表。在取得 Worksheet
# 对象后，可以通过 title 属性取得它的名称。

# 12.3.3 从表中取得单元格
s1 = wb['Sheet1']
print(s1['A1'].value)  # 2015-04-05 13:34:02
type(s1['A1'].value)  # <class 'datetime.datetime'>
c = s1['B1']
print(c.value)  # Apples
'Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value
# 'Row 1, Column 2 is Apples'

'Cell ' + c.coordinate + ' is ' + c.value  # 'Cell B1 is Apples'

sheet['C1'].value  # 73

# Cell 对象有一个 value 属性，不出意外，它包含这个单元格中保存的值。 Cell 对
# 象也有 row、 column 和 coordinate 属性，提供该单元格的位置信息。

# 这里，访问单元格 B1 的 Cell 对象的 value 属性，我们得到字符串'Apples'。 row
# 属性给出的是整数 1， column 属性给出的是'2'， coordinate 属性给出的是'B1'。

# openpyxl 模块将自动解释列 A 中的日期，将它们返回为 datetime 值，而不是字
# 符串。 datetime 数据类型将在第 16 章中进一步解释。


# 用字母来指定列，这在程序中可能有点奇怪，特别是在 Z 列之后，列开时使用
# 两个字母： AA、 AB、 AC 等。作为替代，在调用表的 cell()方法时，可以传入整数
# 作为 row 和 column 关键字参数，也可以得到一个单元格。第一行或第一列的整数
# 是 1，不是 0。输入以下代码，继续交互式环境的例子

cell = sheet.cell(row=1, column=2)
cell.value  # 'Apples'
for i in range(1, 8, 2):
    # 每两行打印一次
    print(i, sheet.cell(row=i, column=2).value)

# 1 Apples
# 3 Pears
# 5 Apples
# 7 Strawberries

# 可以看到，使用表的 cell()方法，传入 row=1 和 column=2，将得到单元格 B1
# 的 Cell 对象，就像指定 sheet['B1']一样。

# 假定你想顺着 B 列，打印出所有奇数行单元格的值。通过传入 2 作为 range()函数
# 的“步长”参数，可以取得每隔一行的单元格（在这里就是所有奇数行）。 for 循环
# 的 i 变量被传递作为 cell()方法的 row 关键字参数，而 column 关键字参数总是取 2。
# 请注意传入的是整数 2，而不是字符串'B'

# 可以通过 Worksheet 对象的 get_highest_row()和 get_highest_column()方法，确定
# 表的大小。在交互式环境中输入以下代码：

sheet = wb['Sheet1']
# sheet.get_highest_row()  # AttributeError: 'Worksheet' object has no attribute 'get_highest_row'
# 获取最大行和最大列
print(sheet.max_row)
print(sheet.max_column)

# 12.3.4 列字母和数字之间的转换
from openpyxl.utils import get_column_letter, column_index_from_string

get_column_letter(1)
get_column_letter(2)
get_column_letter(27)  # 'AA'
get_column_letter(900)  # 'AHP' 26*26+8*26+16=900

sheet = wb['Sheet1']
get_column_letter(sheet.max_column)  # 'C'
column_index_from_string('A')  # 1
column_index_from_string('AA')  # 27

# 可以调用 get_column_letter()， 传入像 27 这样的整数，弄清楚第 27 列的字母是什么。
# 函数 column_index_string()做的事情相反：传入一列的字母名称，它告诉你该列的数字是什么。


# 12.3.5 从表中取得行和列
# 可以将 Worksheet 对象切片，取得电子表格中一行、一列或一个矩形区域中的所有
# Cell 对象。然后可以循环遍历这个切片中的所有单元格。

# 这里，我们指明需要从 A1 到 C3 的矩形区域中的 Cell 对象，得到了一个 Generator
# 对象，它包含该区域中的 Cell 对象。为了帮助我们看清楚这个 Generator 对象，可以
# 使用它的 tuple()方法，在一个元组中列出它的 Cell 对象。

import openpyxl, pprint

sheet = wb['Sheet1']
t = tuple(sheet['A1': 'C3'])
pprint.pprint(t)

# ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>),
#  (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>),
#  (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))

for row_objects in t:
    for o in row_objects:
        print(o.coordinate, o.value)
    print('---END OF ROW---')

# A1 2015-04-05 13:34:02
# B1 Apples
# C1 73
# ---END OF ROW---
# A2 2015-04-05 03:41:23
# B2 Cherries
# C2 85
# ---END OF ROW---
# A3 2015-04-06 12:46:51
# B3 Pears
# C3 14
# ---END OF ROW---

# 要访问特定行或列的单元格的值，也可以利用 Worksheet 对象的 rows 和 columns
# 属性。在交互式环境中输入以下代码
sheet = wb.active
columns = sheet.columns
type(columns)  # <class 'generator'>
l = list(columns)
# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>,
# <Cell 'Sheet1'.A4>, <Cell 'Sheet1'.A5>, <Cell 'Sheet1'.A6>, <Cell 'Sheet1'.A7>)
print(l[0])
print(l[1])
print(l[2])

for cell_o in l[1]:
    print(cell_o.value)

# Since then the behaviour of .columns has changed a little bit -- I'm too lazy to look up exactly when -- and now it
# produces a generator instead (a lazy object which doesn't actually do any work unless it's asked to.)

# 12.3.6 工作簿、工作表、单元格

# 作为快速复习，下面是从电子表格文件中读取单元格涉及的所有函数、方法和
# 数据类型。
# 1．导入 openpyxl 模块。
# 2．调用 openpyxl.load_workbook()函数。
# 3．取得 Workbook 对象。
# 4．调用 wb.active或 get_sheet_by_name()工作簿方法。
# 5．取得 Worksheet 对象。
# 6．使用索引或工作表的 cell()方法，带上 row 和 column 关键字参数。
# 7．取得 Cell 对象。
# 8．读取 Cell 对象的 value 属性。
