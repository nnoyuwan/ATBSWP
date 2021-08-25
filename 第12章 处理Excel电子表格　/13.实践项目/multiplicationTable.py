import sys, openpyxl, os
from openpyxl.utils import get_column_letter, column_index_from_string

n = int(sys.argv[-1])

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, n + 1):
    letter = get_column_letter(i + 1)
    sheet[letter + str(1)] = i
    sheet['A' + str(i + 1)] = i

for j in range(2, sheet.max_column + 1):
    for i in range(2, sheet.max_row + 1):
        sheet[get_column_letter(j) + str(i)] = (i - 1) * (j - 1)

wb.save('multiplicationTable.xlsx')
