import openpyxl, os
from openpyxl.utils import get_column_letter

os.chdir('C:\\Users\\admin\\PycharmProjects\\ATBSWP\\第12章 处理Excel电子表格　\\13.实践项目')

wb = openpyxl.load_workbook('txt2excel.xlsx')
sheet = wb.active

row = sheet.max_row
column = sheet.max_column

n = 1
for j in range(1, column + 1):
    with open('excel2txt%s.txt' % n, 'w') as f:
        for i in range(1, row + 1):
            value = sheet[get_column_letter(j) + str(i)].value
            if value is not None:
                f.write(value)
                f.write('\n')
    n += 1
