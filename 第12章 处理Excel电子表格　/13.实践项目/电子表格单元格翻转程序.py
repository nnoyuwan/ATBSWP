import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

load = openpyxl.load_workbook('myProduce_.xlsx')
load_sheet = load.active

wb = openpyxl.Workbook()
wb_sheet = wb.active

row = load_sheet.max_row
column = load_sheet.max_column

for j in range(1, column + 1):
    for i in range(1, row + 1):
        letter_j = get_column_letter(j)
        letter_i = get_column_letter(i)
        wb_sheet[letter_i + str(j)].value = load_sheet[letter_j + str(i)].value

wb.save('myProduce__.xlsx')
