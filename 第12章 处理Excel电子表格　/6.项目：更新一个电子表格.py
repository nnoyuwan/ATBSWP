# 这个项目需要编写一个程序，更新产品销售电子表格中的单元格。程序将遍
# 历这个电子表格，找到特定类型的产品，并更新它们的价格。


# 每一行代表一次单独的销售。列分别是销售产品的类型（ A）、产品每磅的价格
# （ B）、销售的磅数（ C），以及这次销售的总收入。 TOTAL 列设置为 Excel 公式，将
# 每磅的成本乘以销售的磅数，并将结果取整到分。有了这个公式，如果列 B 或 C 发
# 生变化， TOTAL 列中的单元格将自动更新。


# 现在假设 Garlic、 Celery 和 Lemons 的价格输入的不正确。这让你面对一项无聊
# 的任务：遍历这个电子表格中的几千行，更新所有 garlic、 celery 和 lemon 行中每磅
# 的价格。你不能简单地对价格查找替换，因为可能有其他的产品价格一样，你不希
# 望错误地“更正”。对于几千行数据，手工操作可能要几小时。但你可以编写程序，
# 几秒钟内完成这个任务。

# 你的程序做下面的事情：
# • 循环遍历所有行。
# • 如果该行是 Garlic、 Celery 或 Lemons，更新价格。
# 这意味着代码需要做下面的事情：
# • 打开电子表格文件。
# • 针对每一行，检查列 A 的值是不是 Celery、 Garlic 或 Lemon。
# • 如果是，更新列 B 中的价格。
# • 将该电子表格保存为一个新文件（这样就不会丢失原来的电子表格，以防万一）。


# 第 1 步：利用更新信息建立数据结构
# 需要更新的价格如下：
# Celery 1.19
# Garlic 3.07
# Lemon 1.27
# 你可以像这样编写代码：

# if produceName == 'Celery':
# cellObj = 1.19
# if produceName == 'Garlic':
# cellObj = 3.07
# if produceName == 'Lemon':
# cellObj = 1.27

# 这样硬编码产品和更新的价格有点不优雅。如果你需要用不同的价格，或针对
# 不同的产品，再次更新这个电子表格，就必须修改很多代码。每次修改代码，都有
# 引入缺陷的风险。
# 更灵活的解决方案，是将正确的价格信息保存在字典中，在编写代码时，利用
# 这个数据结构。

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
PRICE_UPDATEs = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27
                 }

for r in range(2, sheet.max_row + 1):
    pName = sheet.cell(row=r, column=1).value
    if pName in PRICE_UPDATEs:
        sheet.cell(row=r, column=2).value = PRICE_UPDATEs[pName]

wb.save('updatedProduceSales.xlsx')
