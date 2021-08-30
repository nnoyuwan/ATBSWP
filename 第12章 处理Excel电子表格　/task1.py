import openpyxl
from openpyxl.worksheet import worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os

os.getcwd()
os.chdir('C:\\Users\\admin\\PycharmProjects\\ATBSWP\\第12章 处理Excel电子表格　')
wb = openpyxl.open('应收账款.21年-4.xlsx', data_only=True)
wb.get_sheet_names()
rec = wb['应收账款']
type(rec)
assert isinstance(rec, openpyxl.worksheet.worksheet.Worksheet)
print(rec.max_row)
print(rec.max_column)

col_index_ = ['C', 'D', 'E', 'G']

# 期初余额， 送货金额， 收款余额， 结余金额
col_index = ['C', 'H', 'I', 'J']
row_index = [4, 51]

c4 = col_index[0] + str(row_index[0])
h51 = col_index[1] + str(row_index[1])
i51 = col_index[2] + str(row_index[1])
j51 = col_index[3] + str(row_index[1])

for n in range(1, 260):  # 遍历259张表
    n_sheet = wb[str(n)]
    # print('********************')
    # print(n_sheet[c4].value)
    # print(n_sheet[h51].value)
    # print(n_sheet[i51].value)
    # print(n_sheet[j51].value)
    # print('********************')
    rec[col_index_[0] + str(n + 2)] = n_sheet[c4].value
    rec[col_index_[1] + str(n + 2)] = n_sheet[h51].value
    rec[col_index_[2] + str(n + 2)] = n_sheet[i51].value
    rec[col_index_[3] + str(n + 2)] = n_sheet[j51].value

    rec[col_index_[0] + str(n + 2)].alignment = Alignment(horizontal='right', vertical='center')
    rec[col_index_[1] + str(n + 2)].alignment = Alignment(horizontal='right', vertical='center')
    rec[col_index_[2] + str(n + 2)].alignment = Alignment(horizontal='right', vertical='center')
    rec[col_index_[3] + str(n + 2)].alignment = Alignment(horizontal='right', vertical='center')

wb.save('应收账款.21年-4_.xlsx')
