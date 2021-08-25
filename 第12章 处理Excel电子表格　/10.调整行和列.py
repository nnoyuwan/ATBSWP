# 在 Excel 中，调整行和列的大小非常容易，只要点击并拖动行的边缘，或列的
# 头部。但如果你需要根据单元格的内容来设置行或列的大小，或者希望设置大量电
# 子表格文件中的行列大小，编写 Python 程序来做就要快得多。

# 行和列也可以完全隐藏起来。或者它们可以“冻结”，这样就总是显示在屏幕
# 上，如果打印该电子表格，它们就出现在每一页上（这很适合做表头

# 12.10.1 设置行高和列宽

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20

wb.save('dimensions.xlsx')

# 12.10.2 合并和拆分单元格
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

# merge_cells()的参数是一个字符串，表示要合并的矩形区域左上角和右下角的
# 单元格： 'A1:D3'将 12 个单元格合并为一个单元格。要设置这些合并后单元格的值，
# 只要设置这一组合并单元格左上角的单元格的值。


# 要拆分单元格，就调用 unmerge_cells()工作表方法。在交互式环境中输入以下
# 代码：

import openpyxl

wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('unmerged.xlsx')

# 12.10.3 冻结窗格
# 对于太大而不能一屏显示的电子表格，“冻结”顶部的几行或最左边的几列，是
# 很有帮助的。例如，冻结的列或行表头，就算用户滚动电子表格，也是始终可见的。
# 这称为“冻结窗格”。在 OpenPyXL 中，每个 Worksheet 对象都有一个 freeze_panes
# 属性，可以设置为一个 Cell 对象或一个单元格坐标的字符串。请注意，单元格上边的所
# 有行和左边的所有列都会冻结，但单元格所在的行和列不会冻结。

# 要解冻所有的单元格，就将 freeze_panes 设置为 None 或'A1'。表 12-3 展示了
# freeze_panes 设定的一些例子，以及哪些行或列会冻结。


# 表 12-3 冻结窗格的例子
# freeze_panes 的设置          冻结的行和列
# sheet.freeze_panes = 'A2'     行 1
# sheet.freeze_panes = 'B1'     列 A
# sheet.freeze_panes = 'C1'     列 A 和列 B
# sheet.freeze_panes = 'C2'     行 1 和列 A 和列 B

# sheet.freeze_panes = 'A1'或    没有冻结窗格
# sheet.freeze_panes = None

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')

# 12.10.4 图表
# openpyxl 支持利用工作表中单元格的数据，创建条形图、折线图、散点图和饼
# 图。要创建图表，需要做下列事情：
# 1．从一个矩形区域选择的单元格，创建一个 Reference 对象。
# 2．通过传入 Reference 对象，创建一个 Series 对象。
# 3．创建一个 Chart 对象。
# 4．将 Series 对象添加到 Chart 对象。
# 5． 可选地设置 Chart 对象的 drawing.top、 drawing.left、 drawing.width 和 drawing.height
# 变量。
# 6．将 Chart 对象添加到 Worksheet 对象。

# Reference 对象需要一些解释。 Reference 对象是通过调用 openpyxl.charts. Reference()
# 函数并传入 3 个参数创建的：
# 1．包含图表数据的 Worksheet 对象。
# 2．两个整数的元组，代表矩形选择区域的左上角单元格，该区域包含图表数
# 据：元组中第一个整数是行，第二个整数是列。请注意第一行是 1，不是 0。
# 3．两个整数的元组，代表矩形选择区域的右下角单元格，该区域包含图表数
# 据：元组中第一个整数是行，第二个整数是列。

from openpyxl.chart import BarChart, ScatterChart, LineChart, LineChart3D, PieChart, Reference, Series

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
    sheet['A' + str(i)] = i

refObj = Reference(sheet, min_col=1, max_col=1, min_row=1, max_row=10)  # 去第1列的1~10行作为y值，x默认从1递

seriesObj = Series(refObj, title='First series')

chartObj = BarChart()
# chartObj = ScatterChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')

from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

wb = Workbook()
ws = wb.active

rows = [
    ['Size', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 25],
    [6, 25, 35],
    [7, 20, 40],
]

for row in rows:
    ws.append(row)

chart = ScatterChart()
chart.title = "Scatter Chart"
chart.style = 13
chart.x_axis.title = 'Size'
chart.y_axis.title = 'Percentage'

xvalues = Reference(ws, min_col=1, min_row=2, max_row=7)  # 取第1列的2~7行作为x
# 加两条线
for i in range(2, 4):
    values = Reference(ws, min_col=i, min_row=1, max_row=7)  # 取第2、3列的2~7行作为y
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)

ws.add_chart(chart, "A10")

wb.save("scatter.xlsx")
