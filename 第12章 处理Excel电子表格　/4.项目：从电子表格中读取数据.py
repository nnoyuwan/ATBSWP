# 假定你有一张电子表格的数据，来自于 2010 年美国人口普查。你有一个无聊的任
# 务，要遍历表中的几千行，计算总的人口，以及每个县的普查区的数目（普查区就是一
# 个地理区域，是为人口普查而定义的）。每行表示一个人口普查区。


# 尽管 Excel 是要能够计算多个选中单元格的和， 你仍然需要选中 3000 个以上县
# 的单元格。即使手工计算一个县的人口只需要几秒钟，整张电子表格也需要几个小
# 时时间

# 在这个项目中，你要编写一个脚本，从人口普查电子表格文件中读取数据，并
# 在几秒钟内计算出每个县的统计值。
# 下面是程序要做的事：
# • 从 Excel 电子表格中读取数据。
# • 计算每个县中普查区的数目。
# • 计算每个县的总人口。
# • 打印结果。
# 这意味着代码需要完成下列任务：
# • 用 openpyxl 模块打开 Excel 文档并读取单元格。
# • 计算所有普查区和人口数据，将它保存到一个数据结构中。
# • 利用 pprint 模块，将该数据结构写入一个扩展名为.py 的文本文件。
import pprint

import openpyxl

# *****第 1 步：读取电子表格数据*****
# censuspopdata.xlsx 电子表格中只有一张表，名为'Population by Census Tract'。每
# 一行都保存了一个普查区的数据。列分别是普查区的编号（ A），州的简称（ B）， 县
# 的名称（ C），普查区的人口（ D）。
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.active
print(sheet.title)  # sheet.title
print(sheet.max_row)  # 72865
countyData = {}

# 保存在 countyData 中的数据结构将是一个字典，以州的简称作为键。每个州的
# 简称将映射到另一个字典，其中的键是该州的县的名称。每个县的名称又映射到一
# 个字典，该字典只有两个键， 'tracts'和'pop'。这些键映射到普查区数目和该县的人口。
# 例如，该字典可能类似于：

# {'AK': {'Aleutians East': {'pop': 3141, 'tracts': 1},
# 'Aleutians West': {'pop': 5561, 'tracts': 2},
# 'Anchorage': {'pop': 291826, 'tracts': 55},
# 'Bethel': {'pop': 17013, 'tracts': 3},
# 'Bristol Bay': {'pop': 997, 'tracts': 1},
# --snip

# 如果前面的字典保存在 countyData 中，下面的表达式求值结果如下：
# >>> countyData['AK']['Anchorage']['pop']
# 291826
# >>> countyData['AK']['Anchorage']['tracts']
# 55

# 一般来说， countyData 字典中的键看起来像这样：
# countyData[state abbrev][county]['tracts']
# countyData[state abbrev][county]['pop']

# 既然知道了 countyData 的结构，就可以编写代码，用县的数据填充它。将下面
# 的代码添加到程序的末尾：
for row in range(2, sheet.max_row + 1):
    State = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # *****第 2 步：填充数据结构*****
    # Make sure the key for this state exists.
    countyData.setdefault(State, {})
    # Make sure the key for this county in this state exists.
    countyData[State].setdefault(county, {'tracts': 0, 'pop': 0})

    # Each row represents one census tract, so increment by one.
    countyData[State][county]['tracts'] += 1
    # Increase the county pop by the pop in this census tract.
    countyData[State][county]['pop'] += int(pop)

print(countyData)

# *****第 3 步：将结果写入文件*****
# Open a new text file and write the contents of countyData to it.
print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData=' + pprint.pformat(countyData))
resultFile.close()
print('Done.')

# pprint.pformat()函数产生一个字符串，它本身就是格式化好的、有效的 Python
# 代码。 将它输出到文本文件 census2010.py，你就通过 Python 程序生成了一个 Python
# 程序！

import os
import sys

os.getcwd()
# os.chdir('C:\\Users\\admin\\PycharmProjects\\ATBSWP\\第12章 处理Excel电子表格　')
# 系统路径要加入cwd才能导包
sys.path.append(os.getcwd())

import census2010

print(census2010.allData['AK']['Anchorage'])

anchoragePop = census2010.allData['AK']['Anchorage']['pop']
print('The 2010 population of Anchorage was ' + str(anchoragePop))

# 第 4 步：类似程序的思想
# 许多公司和组织机构使用 Excel 来保存各种类型的数据，电子表格会变得庞大，这
# 并不少见。解析 Excel 电子表格的程序都有类似的结构：它加载电子表格文件，准备一
# 些变量或数据结构，然后循环遍历电子表格中的每一行。这样的程序可以做下列事情：
# • 比较一个电子表格中多行的数据。
# • 打开多个 Excel 文件，跨电子表格比较数据。
# • 检查电子表格是否有空行或无效的数据，如果有就警告。
# • 从电子表格中读取数据，将它作为 Python 程序的输入。
